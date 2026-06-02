# from rest_framework.decorators import api_view, permission_classes
# from rest_framework.response import Response
# from django.shortcuts import get_object_or_404
# from django.utils import timezone
# from datetime import date
# from io import BytesIO
# import zipfile
# from uuid import uuid4

# from apps.accounts.permissions import IsHR, IsAdmin, IsEmployee
# from apps.employees.models import Employee
# from django.http import HttpResponse
# from django.core.cache import cache

# from .models import Salary, Payslip, PayslipEmailLog, PayrollMonth
# from .serializers import SalarySerializer, PayslipSerializer
# from .services.payroll_calculator import calculate_monthly_salary
# from .utils_pdf import generate_payslip_pdf, generate_payslip_pdf_buffer
# from .tasks import send_payslip_email_task

# from django.conf import settings
# import os
# from reportlab.platypus import Image

# from django.http import HttpResponse
# from django.conf import settings
# from reportlab.platypus import (
#     SimpleDocTemplate,
#     Paragraph,
#     Spacer,
#     Table,
#     TableStyle,
#     Image,
# )
# from reportlab.lib import colors
# from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
# from reportlab.lib.units import inch
# from reportlab.lib.pagesizes import A4
# from reportlab.pdfgen import canvas
# from datetime import datetime
# from decimal import Decimal
# from io import BytesIO
# import os


# # =====================================================
# # SALARY MANAGEMENT
# # =====================================================

# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def set_salary(request):
#     serializer = SalarySerializer(data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data, status=201)
#     return Response(serializer.errors, status=400)


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def all_salaries(request):
#     salaries = Salary.objects.select_related("employee", "employee__user")
#     return Response(SalarySerializer(salaries, many=True).data)


# @api_view(["PUT"])
# @permission_classes([IsHR | IsAdmin])
# def update_salary(request, salary_id):
#     salary = get_object_or_404(Salary, id=salary_id)
#     serializer = SalarySerializer(salary, data=request.data, partial=True)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data)
#     return Response(serializer.errors, status=400)


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def get_salary_by_employee(request, employee_id):
#     try:
#         salary = Salary.objects.get(employee_id=employee_id)
#     except Salary.DoesNotExist:
#         return Response({"error": "Salary not set"}, status=404)
#     return Response(SalarySerializer(salary).data)


# # =====================================================
# # GENERATE PAYSLIP
# # =====================================================

# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def generate_payslip(request):

#     employee_id = request.data.get("employee_id")
#     month_str = request.data.get("month")

#     if not employee_id or not month_str:
#         return Response({"error": "employee_id and month required"}, status=400)

#     year, month_num = map(int, month_str.split("-"))
#     month_date = date(year, month_num, 1)

#     # Block if month closed
#     if PayrollMonth.objects.filter(year=year, month=month_num, status="CLOSED").exists():
#         return Response({"error": "Payroll already closed"}, status=400)

#     employee = get_object_or_404(Employee, id=employee_id)
#     salary = get_object_or_404(Salary, employee=employee)

#     if Payslip.objects.filter(employee=employee, month=month_date).exists():
#         return Response({"error": "Payslip already exists"}, status=400)

#     payroll_data = calculate_monthly_salary(
#         employee=employee,
#         year=year,
#         month=month_num
#     )

#     payslip = Payslip.objects.create(
#         employee=employee,
#         month=month_date,
#         basic=salary.basic,
#         hra=salary.hra,
#         allowances=salary.allowances,
#         fixed_deductions=salary.deductions,
#         gross_salary=payroll_data["gross_salary"],
#         working_days=payroll_data["working_days"],
#         absent_days=payroll_data["absent_days"],
#         unpaid_leave_days=payroll_data["unpaid_leave_days"],
#         half_days=payroll_data["half_days"],
#         late_days=payroll_data["late_days"],
#         attendance_deduction=payroll_data["attendance_deduction"],
#         late_penalty=payroll_data["late_penalty"],
#         net_pay=payroll_data["net_salary"],
#         status="DRAFT"
#     )

#     return Response({
#         "message": "Payslip generated",
#         "payslip": PayslipSerializer(payslip).data
#     }, status=201)


# # =====================================================
# # BULK GENERATE
# # =====================================================

# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def bulk_generate_payslips(request):

#     month_str = request.data.get("month")
#     if not month_str:
#         return Response({"error": "Month required"}, status=400)

#     year, month_num = map(int, month_str.split("-"))
#     month_date = date(year, month_num, 1)

#     if PayrollMonth.objects.filter(year=year, month=month_num, status="CLOSED").exists():
#         return Response({"error": "Payroll already closed"}, status=400)

#     employees = Employee.objects.all()
#     generated = 0

#     for emp in employees:
#         try:
#             salary = Salary.objects.get(employee=emp)
#         except Salary.DoesNotExist:
#             continue

#         if Payslip.objects.filter(employee=emp, month=month_date).exists():
#             continue

#         payroll_data = calculate_monthly_salary(emp, year, month_num)

#         Payslip.objects.create(
#             employee=emp,
#             month=month_date,
#             basic=salary.basic,
#             hra=salary.hra,
#             allowances=salary.allowances,
#             fixed_deductions=salary.deductions,
#             gross_salary=payroll_data["gross_salary"],
#             working_days=payroll_data["working_days"],
#             absent_days=payroll_data["absent_days"],
#             unpaid_leave_days=payroll_data["unpaid_leave_days"],
#             half_days=payroll_data["half_days"],
#             late_days=payroll_data["late_days"],
#             attendance_deduction=payroll_data["attendance_deduction"],
#             late_penalty=payroll_data["late_penalty"],
#             net_pay=payroll_data["net_salary"],
#             status="DRAFT"
#         )

#         generated += 1

#     return Response({"generated": generated})


# # =====================================================
# # APPROVE & CLOSE
# # =====================================================

# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def bulk_approve_payslips(request):

#     month_str = request.data.get("month")
#     year, month_num = map(int, month_str.split("-"))
#     month_date = date(year, month_num, 1)

#     payslips = Payslip.objects.filter(month=month_date, status="DRAFT")
#     count = payslips.count()
#     payslips.update(status="APPROVED")

#     PayrollMonth.objects.update_or_create(
#         year=year,
#         month=month_num,
#         defaults={"status": "CLOSED", "closed_on": timezone.now()}
#     )

#     return Response({"approved_count": count})


# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def approve_payslip(request, payslip_id):
#     payslip = get_object_or_404(Payslip, id=payslip_id)
#     payslip.status = "APPROVED"
#     payslip.save()
#     return Response({"message": "Approved"})


# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def mark_payslip_paid(request, payslip_id):
#     payslip = get_object_or_404(Payslip, id=payslip_id)
#     payslip.status = "PAID"
#     payslip.paid_on = timezone.now()
#     payslip.save()
#     return Response({"message": "Paid"})


# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def cancel_payslip(request, payslip_id):
#     payslip = get_object_or_404(Payslip, id=payslip_id)
#     payslip.status = "CANCELLED"
#     payslip.save()
#     return Response({"message": "Cancelled"})


# # =====================================================
# # VIEW PAYSLIPS
# # =====================================================

# @api_view(["GET"])
# @permission_classes([IsEmployee])
# def my_payslips(request):
#     employee = request.user.employee_profile
#     slips = Payslip.objects.filter(employee=employee)
#     return Response(PayslipSerializer(slips, many=True).data)


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def all_payslips(request):
#     slips = Payslip.objects.all()
#     return Response(PayslipSerializer(slips, many=True).data)


# # =====================================================
# # PDF & ZIP DOWNLOAD
# # =====================================================

# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin | IsEmployee])
# def download_payslip_pdf(request, payslip_id):
#     payslip = get_object_or_404(Payslip, id=payslip_id)
#     return generate_payslip_pdf(payslip)


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def download_all_payslips_zip(request):

#     month_str = request.query_params.get("month")
#     year, month_num = map(int, month_str.split("-"))
#     month_date = date(year, month_num, 1)

#     payslips = Payslip.objects.filter(month=month_date)

#     zip_buffer = BytesIO()
#     with zipfile.ZipFile(zip_buffer, "w") as zip_file:
#         for slip in payslips:
#             pdf_buffer = generate_payslip_pdf_buffer(slip)
#             filename = f"{slip.employee.employee_id}_{month_str}.pdf"
#             zip_file.writestr(filename, pdf_buffer.getvalue())

#     zip_buffer.seek(0)
#     response = HttpResponse(zip_buffer, content_type="application/zip")
#     response["Content-Disposition"] = f'attachment; filename="Payslips_{month_str}.zip"'
#     return response


# # =====================================================
# # EMAIL SYSTEM
# # =====================================================

# @api_view(["POST"])
# @permission_classes([IsHR | IsAdmin])
# def bulk_email_payslips(request):

#     month_str = request.data.get("month")
#     year, month_num = map(int, month_str.split("-"))
#     month_date = date(year, month_num, 1)

#     payslips = Payslip.objects.filter(month=month_date, status="APPROVED")

#     batch_id = str(uuid4())
#     cache.set(f"bulk_{batch_id}_total", payslips.count())
#     cache.set(f"bulk_{batch_id}_completed", 0)
#     cache.set(f"bulk_{batch_id}_failed", 0)

#     for slip in payslips:
#         send_payslip_email_task.delay(slip.id, batch_id)

#     return Response({"batch_id": batch_id})


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def bulk_email_progress(request, batch_id):
#     return Response({
#         "total": cache.get(f"bulk_{batch_id}_total", 0),
#         "completed": cache.get(f"bulk_{batch_id}_completed", 0),
#         "failed": cache.get(f"bulk_{batch_id}_failed", 0)
#     })


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def payslip_email_logs(request, payslip_id):
#     logs = PayslipEmailLog.objects.filter(payslip_id=payslip_id)
#     return Response([
#         {
#             "email": log.email,
#             "status": log.status,
#             "sent_at": log.sent_at,
#             "error": log.error_message
#         }
#         for log in logs
#     ])


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def email_dashboard(request):

#     month_str = request.query_params.get("month")
#     year, month_num = map(int, month_str.split("-"))
#     month_date = date(year, month_num, 1)

#     logs = PayslipEmailLog.objects.filter(payslip__month=month_date)

#     total = logs.count()
#     success = logs.filter(status="SUCCESS").count()
#     failed = logs.filter(status="FAILED").count()
#     pending = logs.filter(status="PENDING").count()

#     return Response({
#         "total": total,
#         "success": success,
#         "failed": failed,
#         "pending": pending
#     })

# # =====================================================
# # PAYROLL STATUS DASHBOARD
# # =====================================================

# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def payroll_status(request):

#     month_str = request.query_params.get("month")
#     status_filter = request.query_params.get("status")

#     if not month_str:
#         return Response({"error": "Month required"}, status=400)

#     try:
#         year, month_num = map(int, month_str.split("-"))
#         month_date = date(year, month_num, 1)
#     except ValueError:
#         return Response({"error": "Invalid month format"}, status=400)

#     employees = Employee.objects.select_related("user").all()
#     data = []

#     for emp in employees:
#         salary_exists = Salary.objects.filter(employee=emp).exists()

#         payslip = Payslip.objects.filter(
#             employee=emp,
#             month=month_date
#         ).first()

#         if status_filter and status_filter != "ALL":
#             if not payslip or payslip.status != status_filter:
#                 continue

#         data.append({
#             "employee_id": emp.id,
#             "employee_name": f"{emp.first_name} {emp.last_name}",
#             "salary_set": salary_exists,
#             "payslip_generated": bool(payslip),
#             "payslip_status": payslip.status if payslip else None,
#             "payslip_id": payslip.id if payslip else None,
#         })

#     return Response(data)


# # def generate_payslip_pdf(payslip):

# #     response = HttpResponse(content_type="application/pdf")
# #     filename = f"Payslip_{payslip.employee.employee_id}_{payslip.month}.pdf"
# #     response["Content-Disposition"] = f'attachment; filename="{filename}"'

# #     doc = SimpleDocTemplate(
# #         response,
# #         pagesize=A4,
# #         rightMargin=30,
# #         leftMargin=30,
# #         topMargin=30,
# #         bottomMargin=30,
# #     )

# #     elements = []
# #     styles = getSampleStyleSheet()

# #     # =====================================================
# #     # 🔥 SAFE COMPANY LOGO BLOCK (ADD HERE)
# #     # =====================================================

# #     try:
# #         logo_path = os.path.join(settings.BASE_DIR, "static/company/logo.png")

# #         if os.path.isfile(logo_path):
# #             logo = Image(logo_path, width=120, height=60)
# #             logo.hAlign = "CENTER"
# #             elements.append(logo)
# #             elements.append(Spacer(1, 0.2 * inch))

# #     except Exception as e:
# #         print("Logo load error:", e)

# #     # =====================================================
# #     # COMPANY HEADER
# #     # =====================================================

# #     header_style = styles["Heading1"]
# #     header_style.alignment = 1

# #     elements.append(Paragraph("GMMC HRMS", header_style))
# #     elements.append(Spacer(1, 0.2 * inch))



# def generate_payslip_pdf(payslip):

#     response = HttpResponse(content_type="application/pdf")
#     filename = f"Payslip_{payslip.employee.employee_id}_{payslip.month}.pdf"
#     response["Content-Disposition"] = f'attachment; filename="{filename}"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=A4,
#         rightMargin=30,
#         leftMargin=30,
#         topMargin=30,
#         bottomMargin=30,
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # =====================================================
#     # SAFE COMPANY LOGO
#     # =====================================================

#     try:
#         logo_path = os.path.join(settings.BASE_DIR, "static/company/logo.png")

#         if os.path.isfile(logo_path):
#             logo = Image(logo_path, width=120, height=60)
#             logo.hAlign = "CENTER"
#             elements.append(logo)
#             elements.append(Spacer(1, 0.2 * inch))

#     except Exception as e:
#         print("Logo load error:", e)

#     # =====================================================
#     # COMPANY HEADER
#     # =====================================================

#     header_style = styles["Heading1"]
#     header_style.alignment = 1

#     elements.append(Paragraph("GMMC HRMS", header_style))
#     elements.append(Spacer(1, 0.2 * inch))

#     elements.append(Paragraph("Salary Payslip", styles["Heading2"]))
#     elements.append(Spacer(1, 0.3 * inch))

#     # =====================================================
#     # EMPLOYEE INFO TABLE
#     # =====================================================

#     employee_data = [
#         ["Employee Name", payslip.employee.user.username],
#         ["Employee ID", payslip.employee.employee_id],
#         ["Month", payslip.month.strftime("%B %Y")],
#         ["Generated On", datetime.now().strftime("%d-%m-%Y %H:%M")],
#     ]

#     emp_table = Table(employee_data, colWidths=[150, 300])
#     emp_table.setStyle(
#         TableStyle([
#             ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
#             ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
#             ("FONTSIZE", (0, 0), (-1, -1), 10),
#             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
#         ])
#     )

#     elements.append(emp_table)
#     elements.append(Spacer(1, 0.4 * inch))

#     # =====================================================
#     # SALARY BREAKDOWN
#     # =====================================================

#     salary_data = [
#         ["Component", "Amount (₹)"],
#         ["Basic", f"{payslip.basic:.2f}"],
#         ["HRA", f"{payslip.hra:.2f}"],
#         ["Allowances", f"{payslip.allowances:.2f}"],
#         ["Gross Salary", f"{payslip.gross_salary:.2f}"],
#         ["Attendance Deduction", f"{payslip.attendance_deduction:.2f}"],
#         ["Late Penalty", f"{payslip.late_penalty:.2f}"],
#         ["Fixed Deductions", f"{payslip.fixed_deductions:.2f}"],
#         ["Net Pay", f"{payslip.net_pay:.2f}"],
#     ]

#     salary_table = Table(salary_data, colWidths=[250, 200])
#     salary_table.setStyle(
#         TableStyle([
#             ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
#             ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
#             ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
#         ])
#     )

#     elements.append(salary_table)
#     elements.append(Spacer(1, 0.5 * inch))

#     # =====================================================
#     # FOOTER
#     # =====================================================

#     footer_style = ParagraphStyle(
#         name="FooterStyle",
#         parent=styles["Normal"],
#         fontSize=8,
#         textColor=colors.grey,
#         alignment=1,
#     )

#     elements.append(
#         Paragraph(
#             "This is a system generated payslip and does not require signature.",
#             footer_style,
#         )
#     )

#     # =====================================================
#     # BUILD PDF
#     # =====================================================

#     doc.build(elements)

#     return response


# @api_view(["POST"])
# @permission_classes([IsAdmin])  # 🔐 Only Admin
# def reopen_payroll_month(request):

#     month_str = request.data.get("month")

#     if not month_str:
#         return Response({"error": "Month required"}, status=400)

#     try:
#         year, month_num = map(int, month_str.split("-"))
#     except ValueError:
#         return Response({"error": "Invalid month format"}, status=400)

#     payroll_month = PayrollMonth.objects.filter(
#         year=year,
#         month=month_num,
#         status="CLOSED"
#     ).first()

#     if not payroll_month:
#         return Response({"error": "Month is not closed"}, status=400)

#     payroll_month.status = "OPEN"
#     payroll_month.closed_on = None
#     payroll_month.save()

#     return Response({
#         "message": "Payroll month reopened successfully"
#     })


# @api_view(["GET"])
# @permission_classes([IsHR | IsAdmin])
# def payroll_dashboard_summary(request):

#     from datetime import date

#     today = date.today()
#     year = today.year
#     month_num = today.month
#     month_date = date(year, month_num, 1)

#     from .models import PayrollMonth

#     payroll_month = PayrollMonth.objects.filter(
#         year=year,
#         month=month_num
#     ).first()

#     payroll_status = "OPEN"
#     if payroll_month and payroll_month.status == "CLOSED":
#         payroll_status = "CLOSED"

#     payslips = Payslip.objects.filter(month=month_date)

#     total = payslips.count()
#     draft = payslips.filter(status="DRAFT").count()
#     approved = payslips.filter(status="APPROVED").count()
#     paid = payslips.filter(status="PAID").count()

#     return Response({
#         "month": month_date.strftime("%B %Y"),
#         "status": payroll_status,
#         "total": total,
#         "draft": draft,
#         "approved": approved,
#         "paid": paid
#     })


# ============================================================
# ENTERPRISE PAYROLL MODULE (Celery-Ready Architecture)
# ============================================================

import csv
from django.http import HttpResponse
from decimal import Decimal
from datetime import datetime, date
from decimal import Decimal
import io
import zipfile
from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django.core.mail import EmailMessage
from django.db.models import Sum
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from apps.accounts.permissions import IsHR, IsEmployee
from apps.employees.models import Employee
from .models import Salary, Payslip, PayslipEmailLog, PayrollMonth
from calendar import monthrange
from apps.attendance.models import Attendance
from apps.leaves.models import LeaveRequest
from .models import ProfessionalTaxSlab
from apps.leaves.models import LeaveBalance
from django.http import HttpResponse
from decimal import Decimal, ROUND_HALF_UP
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from django.db import transaction
from decimal import Decimal
from datetime import datetime
from django.db import transaction
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from apps.accounts.permissions import IsHR
from apps.employees.models import Employee
from .models import Payslip
from .tax_engine import calculate_monthly_tds
from datetime import datetime
from decimal import Decimal
from calendar import monthrange
from django.db import transaction
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status

from apps.accounts.permissions import IsHR
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from apps.accounts.permissions import IsHR
from decimal import Decimal
from apps.employees.models import Employee
from django.db.models import Sum
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from apps.accounts.permissions import IsEmployee
# from .models import PayrollRecord
from django.db.models import Sum
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from apps.accounts.permissions import IsEmployee
from .models import Payslip
from calendar import monthrange
from decimal import Decimal
from apps.payroll.services.lop_service import calculate_lop_for_month
from django.http import HttpResponse
from .utils.payslip_pdf import generate_payslip_pdf
from .models import Payslip
from rest_framework.decorators import api_view, permission_classes
from apps.accounts.permissions import IsEmployee
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db.models import Sum
from rest_framework.response import Response
from django.db.models import Sum, Avg, Count
from .models import Salary
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from .models import Salary
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import pagesizes
from django.db.models import Sum, Avg
from datetime import datetime
from .models import Payslip
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import pagesizes
from reportlab.lib.units import inch
from django.db.models import Sum
from django.http import HttpResponse
from datetime import datetime
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from .models import Payslip
from .utils.salary_utils import get_current_salary
from .services.payroll_service import (
    PAYROLL_COMPONENT_FIELDS,
    build_salary_record,
    normalize_salary_data,
)
from .utils.payroll_calculations import calculate_salary_totals, calculate_payslip_totals, to_decimal
from rest_framework.viewsets import ModelViewSet
from .models import SalaryRevision
from .serializers import EmployeeSalarySerializer, SalaryRevisionSerializer, SalarySerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from decimal import Decimal
from .models import SalaryRevision
from apps.employees.models import Employee
from apps.accounts.tenant_utils import get_current_company
from apps.audit.utils import log_action
from apps.payroll.models import Payslip, PayrollMonth
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Sum
from django.utils import timezone
from .utils.payslip_pdf import generate_payslip_pdf


# ============================================================
# SALARY MANAGEMENT
# ============================================================

# ============================================================
# SALARY MANAGEMENT (MULTI-COMPONENT)
# ============================================================


def _build_salary_defaults(payload, existing_salary=None):
    if existing_salary is None:
        return build_salary_record(payload)

    merged_payload = {
        field_name: payload.get(field_name, getattr(existing_salary, field_name, 0))
        for field_name in PAYROLL_COMPONENT_FIELDS
    }
    return build_salary_record(merged_payload)

@api_view(["POST"])
@permission_classes([IsHR])
def set_salary(request):
    employee_id = request.data.get("employee_id")
    company = get_current_company(request)

    if not employee_id:
        return Response({"error": "employee_id required"}, status=400)

    qs = Employee.objects.filter(id=employee_id)
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    try:
        employee = qs.get()
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    # 🔒 STABILITY CHECK — Prevent salary change after payslip exists
    existing_payslips = Payslip.objects.filter(employee=employee)

    if existing_payslips.exists():
        return Response(
            {"error": "Cannot modify salary after payslip generation"},
            status=400
        )

    # ✅ Safe to update (include company for tenant)
    defaults = _build_salary_defaults(request.data)
    if getattr(employee, "company_id", None):
        defaults["company_id"] = employee.company_id
    salary, created = Salary.objects.update_or_create(
        employee=employee,
        defaults=defaults,
    )

    return Response({
        "message": "Salary structure saved successfully",
        "created": created,
        "salary": EmployeeSalarySerializer(salary).data,
    }, status=201 if created else 200)


@api_view(["GET"])
@permission_classes([IsHR])
def all_salaries(request):
    company = get_current_company(request)
    qs = Salary.objects.select_related("employee")
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    return Response(SalarySerializer(qs, many=True).data)


@api_view(["PUT"])
@permission_classes([IsHR])
def update_salary(request, salary_id):
    company = get_current_company(request)
    qs = Salary.objects.filter(id=salary_id)
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    try:
        salary = qs.get()
    except Salary.DoesNotExist:
        return Response({"error": "Salary not found"}, status=404)

    salary_defaults = _build_salary_defaults(request.data, existing_salary=salary)
    for field_name, value in salary_defaults.items():
        setattr(salary, field_name, value)
    salary.save()

    return Response({
        "message": "Salary updated successfully",
        "salary": EmployeeSalarySerializer(salary).data,
    })


@api_view(["GET"])
@permission_classes([IsHR])
def get_salary_by_employee(request, employee_id):
    company = get_current_company(request)
    qs = Salary.objects.filter(employee__id=employee_id)
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    try:
        salary = qs.get()
    except Salary.DoesNotExist:
        return Response({"error": "Salary not found"}, status=404)
    return Response(EmployeeSalarySerializer(salary).data)


def calculate_epf(basic_after_lop, employee):
    if not getattr(employee, 'pf_applicable', False):
        return Decimal("0.00"), Decimal("0.00")

    PF_RATE = Decimal("0.12")
    PF_WAGE_CEILING = Decimal("15000")

    # Apply ceiling (default to True if not set)
    pf_wage_ceiling_applicable = getattr(employee, 'pf_wage_ceiling_applicable', True)
    if pf_wage_ceiling_applicable:
        pf_wage = min(basic_after_lop, PF_WAGE_CEILING)
    else:
        pf_wage = basic_after_lop

    employee_pf = pf_wage * PF_RATE
    employer_pf = pf_wage * PF_RATE

    return employee_pf, employer_pf


def calculate_esi(gross_after_lop, employee):
    if not getattr(employee, 'esi_applicable', False):
        return Decimal("0.00"), Decimal("0.00")

    ESI_EMPLOYEE_RATE = Decimal("0.0075")
    ESI_EMPLOYER_RATE = Decimal("0.0325")
    ESI_LIMIT = Decimal("21000")

    if gross_after_lop > ESI_LIMIT:
        return Decimal("0.00"), Decimal("0.00")

    employee_esi = gross_after_lop * ESI_EMPLOYEE_RATE
    employer_esi = gross_after_lop * ESI_EMPLOYER_RATE

    return employee_esi, employer_esi

# ============================================================
# PAYSLIP GENERATION (Enterprise Indian Payroll Version)
# ============================================================

@api_view(["POST"])
@permission_classes([IsHR])
@transaction.atomic
def generate_payslip(request):
    company = get_current_company(request)
    employee_id = request.data.get("employee_id")
    month_value = request.data.get("month")

    if not employee_id or not month_value:
        return Response({"error": "employee_id and month required"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m").date()
        year = month_date.year
        month = month_date.month
    except ValueError:
        return Response({"error": "Invalid month format (YYYY-MM)"}, status=400)

    defaults_pm = {"status": "OPEN"}
    pm_kwargs = {"year": year, "month": month}
    if company and hasattr(PayrollMonth, "company"):
        defaults_pm["company"] = company
        pm_kwargs["company"] = company
    payroll_month, created = PayrollMonth.objects.get_or_create(
        defaults=defaults_pm, **pm_kwargs
    )
    if payroll_month.status == "CLOSED":
        payroll_month.status = "OPEN"
        payroll_month.save(update_fields=["status"])

    qs = Employee.objects.select_related("salary").filter(id=employee_id)
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    try:
        employee = qs.get()
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    if not employee.is_active:
        return Response({"error": "Cannot generate payslip for inactive employee"}, status=400)

    salary = get_current_salary(employee)
    if not salary:
        return Response({"error": "Salary structure not set"}, status=400)

    salary_totals = calculate_salary_totals(salary)
    if salary_totals["gross_salary"] <= 0:
        return Response(
            {"error": "Invalid salary structure. Gross salary must be greater than 0."},
            status=400,
        )

    if Payslip.objects.filter(employee=employee, month__year=year, month__month=month).exists():
        return Response({"error": "Payslip already generated"}, status=400)

    lop_days = calculate_lop_for_month(employee, year, month)
    payslip_totals = calculate_payslip_totals(
        salary_obj=salary,
        year=year,
        month=month,
        lop_days=lop_days,
    )

    payslip_kw = {
        "employee": employee,
        "month": month_date,
        "basic": payslip_totals["basic"],
        "da": payslip_totals["da"],
        "hra": payslip_totals["hra"],
        "conveyance": payslip_totals["conveyance"],
        "medical": payslip_totals["medical"],
        "special_allowance": payslip_totals["special_allowance"],
        "gross_salary": payslip_totals["gross_salary"],
        "lop_days": payslip_totals["lop_days"],
        "lop_deduction": payslip_totals["lop_deduction"],
        "employee_pf": payslip_totals["employee_pf"],
        "employer_pf": payslip_totals["employer_pf"],
        "employee_esi": payslip_totals["employee_esi"],
        "employer_esi": payslip_totals["employer_esi"],
        "professional_tax": payslip_totals["professional_tax"],
        "tds_amount": payslip_totals["tds"],
        "medical_insurance": payslip_totals["medical_insurance"],
        "fixed_deductions": payslip_totals["base_deductions"],
        "net_pay": payslip_totals["net_pay"],
        "status": "NOT PAID",
    }
    if getattr(employee, "company_id", None):
        payslip_kw["company_id"] = employee.company_id
    payslip = Payslip.objects.create(**payslip_kw)

    log_action(
        request, "GENERATE", "Payslip",
        object_id=payslip.id,
        description=f"Payslip generated for employee {employee.id}, {month_value}",
        company=company,
    )
    return Response(
        {
            "message": "Payslip generated successfully",
            "payroll_month_created": created,
            "payslip_id": payslip.id,
            "gross_salary": payslip_totals["gross_salary"],
            "lop_days": payslip_totals["lop_days"],
            "lop_deduction": payslip_totals["lop_deduction"],
            "total_deductions": payslip_totals["total_deductions"],
            "net_pay": payslip_totals["net_pay"],
            "ctc": salary_totals["ctc"],
        },
        status=201,
    )


def calculate_lop(employee, year, month, gross_salary):
    total_days = monthrange(year, month)[1]

    # 1️⃣ Unpaid approved leaves
    unpaid_leaves = LeaveRequest.objects.filter(
        employee=employee,
        status="APPROVED",
        leave_type__is_paid=False,
        start_date__year=year,
        start_date__month=month,
    )

    unpaid_days = 0
    for leave in unpaid_leaves:
        unpaid_days += (leave.end_date - leave.start_date).days + 1

    # 2️⃣ ABSENT attendance
    absent_days = Attendance.objects.filter(
        employee=employee,
        date__year=year,
        date__month=month,
        status="ABSENT"
    ).count()

    lop_days = unpaid_days + absent_days

    per_day_salary = Decimal(gross_salary) / Decimal(total_days)
    lop_deduction = per_day_salary * Decimal(lop_days)

    return lop_days, lop_deduction

@api_view(["POST"])
@permission_classes([IsHR])
def bulk_generate_payslips(request):
    month_value = request.data.get("month")
    if not month_value:
        return Response({"error": "month required"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m").date()
    except ValueError:
        return Response({"error": "Invalid month format (YYYY-MM)"}, status=400)

    company = get_current_company(request)
    pm_defaults = {"status": "OPEN"}
    pm_kwargs = {"year": month_date.year, "month": month_date.month}
    if company and hasattr(PayrollMonth, "company"):
        pm_defaults["company"] = company
        pm_kwargs["company"] = company
    PayrollMonth.objects.get_or_create(
        defaults=pm_defaults, **pm_kwargs
    )

    employees = Employee.objects.filter(is_active=True)
    if company is not None and hasattr(employees.model, 'company'):
        employees = employees.filter(company=company)
    generated = 0

    for emp in employees:
        salary = get_current_salary(emp)
        if not salary:
            continue

        if calculate_salary_totals(salary)["gross_salary"] <= 0:
            continue

        lop_days = calculate_lop_for_month(emp, month_date.year, month_date.month)
        totals = calculate_payslip_totals(
            salary_obj=salary,
            year=month_date.year,
            month=month_date.month,
            lop_days=lop_days,
        )

        payslip_defaults = {
            "basic": totals["basic"],
            "da": totals["da"],
            "hra": totals["hra"],
            "conveyance": totals["conveyance"],
            "medical": totals["medical"],
            "special_allowance": totals["special_allowance"],
            "gross_salary": totals["gross_salary"],
            "lop_days": totals["lop_days"],
            "lop_deduction": totals["lop_deduction"],
            "employee_pf": totals["employee_pf"],
            "employer_pf": totals["employer_pf"],
            "employee_esi": totals["employee_esi"],
            "employer_esi": totals["employer_esi"],
            "professional_tax": totals["professional_tax"],
            "tds_amount": totals["tds"],
            "medical_insurance": totals["medical_insurance"],
            "fixed_deductions": totals["base_deductions"],
            "net_pay": totals["net_pay"],
            "status": "NOT PAID",
        }
        if getattr(emp, "company_id", None):
            payslip_defaults["company_id"] = emp.company_id
        _, created = Payslip.objects.get_or_create(
            employee=emp,
            month=date(month_date.year, month_date.month, 1),
            defaults=payslip_defaults,
        )
        if created:
            generated += 1

    log_action(
        request, "GENERATE", "Payslip",
        description=f"Bulk payroll generated: {generated} payslips for {month_value}",
        company=company,
    )
    return Response({"generated": generated})


# ============================================================
# PAYSLIP STATUS ACTIONS
# ============================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def approve_payslip(request, payslip_id):
    company = get_current_company(request)
    qs = Payslip.objects.select_related("employee").filter(id=payslip_id)
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    try:
        payslip = qs.get()
    except Payslip.DoesNotExist:
        return Response({"error": "Payslip not found"}, status=404)

    # Approve payslip
    payslip.status = "APPROVED"
    payslip.save()

    year = payslip.month.year
    month = payslip.month.month

    # Check if any payslip still NOT approved
    pending = Payslip.objects.filter(
        month__year=year,
        month__month=month
    ).exclude(status="APPROVED").exists()

    return Response({
        "message": "Payslip approved successfully"
    })


@api_view(["POST"])
@permission_classes([IsHR])
def mark_payslip_paid(request, payslip_id):
    company = get_current_company(request)
    qs = Payslip.objects.filter(id=payslip_id)
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    now = timezone.now()
    bank_reference = request.data.get("bank_reference")
    qs.update(
        status="PAID",
        paid_on=now,
    )
    return Response({"message": "Payslip marked as PAID"})


@api_view(["POST"])
@permission_classes([IsHR])
def cancel_payslip(request, payslip_id):
    company = get_current_company(request)
    qs = Payslip.objects.filter(id=payslip_id)
    if company is not None and hasattr(qs.model, 'company'):
        qs = qs.filter(company=company)
    qs.update(status="CANCELLED")
    return Response({"message": "Payslip cancelled"})


# ============================================================
# PAYSLIP PDF (Professional)
# ============================================================

# def generate_payslip_pdf(payslip):

#     buffer = io.BytesIO()
#     doc = SimpleDocTemplate(buffer, pagesize=A4)
#     elements = []
#     styles = getSampleStyleSheet()

#     # =========================================================
#     # HEADER
#     # =========================================================
#     # =========================================================
#     # HEADER
#     # =========================================================
#     elements.append(Paragraph("<b>PAYSLIP</b>", styles["Title"]))
#     elements.append(Spacer(1, 12))

#     company_name = "GMMC Company Pvt Ltd"

#     elements.append(Paragraph(
#         f"Company: {company_name}",
#         styles["Normal"]
#     ))

#     elements.append(Paragraph(
#         f"Month: {payslip.month.strftime('%B %Y')}",
#         styles["Normal"]
#     ))

#     elements.append(Spacer(1, 20))

#     # =========================================================
#     # EMPLOYEE DETAILS
#     # =========================================================
#     employee = payslip.employee

#     total_days = monthrange(payslip.month.year, payslip.month.month)[1]

#     employee_data = [
#         ["EMPCODE", getattr(employee, "employee_id", ""), "PF NO", getattr(employee, "pf_number", "")],
#         ["EMPNAME", getattr(employee, "first_name", ""), "STD DAYS", total_days],
#         ["DESIGNATION", getattr(employee, "designation", ""), "WRK DAYS", total_days],
#         ["DOJ", getattr(employee, "date_of_joining", ""), "LOP DAYS", payslip.lop_days],
#         ["BUSINESS UNIT", getattr(employee, "department", ""), "BANK NAME", getattr(employee, "bank_name", "")],
#         ["PAN", getattr(employee, "pan_number", ""), "ACCOUNT NO", getattr(employee, "bank_account_number", "")],
#         ["LOCATION", getattr(employee, "location", ""), "UAN", getattr(employee, "uan_number", "")],
#     ]

#     table = Table(employee_data, colWidths=[120,150,120,150])

#     emp_table = Table(employee_data, colWidths=[2.5 * inch, 3 * inch])
#     emp_table.setStyle([
#         ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
#     ])

#     elements.append(emp_table)
#     elements.append(Spacer(1, 20))

#     # =========================================================
#     # EARNINGS (A)
#     # =========================================================
#     earnings_data = [
#         ["Component", "Amount (₹)"],
#         ["Basic", payslip.basic],
#         ["DA", payslip.da],
#         ["HRA", payslip.hra],
#         ["Conveyance", payslip.conveyance],
#         ["Medical", payslip.medical],
#         ["Special Allowance", payslip.special_allowance],
#         ["Bonus", payslip.bonus],
#         ["Other Allowance", payslip.other_allowance],
#         ["Gross Salary (A)", payslip.gross_salary],
#     ]

#     earnings_table = Table(earnings_data, colWidths=[3 * inch, 2 * inch])
#     earnings_table.setStyle([
#         ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
#         ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
#     ])

#     elements.append(Paragraph("<b>EARNINGS</b>", styles["Heading3"]))
#     elements.append(Spacer(1, 10))
#     elements.append(earnings_table)
#     elements.append(Spacer(1, 20))

#     # =========================================================
#     # DEDUCTIONS (B)
#     # =========================================================
#     total_deductions = (
#         payslip.lop_deduction +
#         payslip.employee_pf +
#         payslip.employee_esi +
#         payslip.professional_tax +
#         payslip.tds_amount +
#         payslip.medical_insurance
#     )

#     deductions_data = [
#         ["Component", "Amount (₹)"],
#         ["LOP Deduction", payslip.lop_deduction],
#         ["Employee PF", payslip.employee_pf],
#         ["Employee ESI", payslip.employee_esi],
#         ["Professional Tax", payslip.professional_tax],
#         ["TDS", payslip.tds_amount],
#         # ["Fixed Deductions", payslip.fixed_deductions],
#         ["Fixed Deductions", getattr(payslip, "fixed_deductions", 0)],
#         ["Total Deductions (B)", total_deductions],
#     ]

#     deductions_table = Table(deductions_data, colWidths=[3 * inch, 2 * inch])
#     deductions_table.setStyle([
#         ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
#         ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
#     ])

#     elements.append(Paragraph("<b>DEDUCTIONS</b>", styles["Heading3"]))
#     elements.append(Spacer(1, 10))
#     elements.append(deductions_table)
#     elements.append(Spacer(1, 20))

#     # =========================================================
#     # EMPLOYER CONTRIBUTIONS (C)
#     # =========================================================
#     employer_data = [
#         ["Component", "Amount (₹)"],
#         ["Employer PF", payslip.employer_pf],
#         ["Employer ESI", payslip.employer_esi],
#     ]

#     employer_table = Table(employer_data, colWidths=[3 * inch, 2 * inch])
#     employer_table.setStyle([
#         ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
#         ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
#     ])

#     elements.append(Paragraph("<b>EMPLOYER CONTRIBUTIONS</b>", styles["Heading3"]))
#     elements.append(Spacer(1, 10))
#     elements.append(employer_table)
#     elements.append(Spacer(1, 20))

#     # =========================================================
#     # FINAL SUMMARY
#     # =========================================================
#     ctc = payslip.gross_salary + payslip.employer_pf + payslip.employer_esi

#     summary_data = [
#         ["Net Salary (A - B)", payslip.net_pay],
#         ["CTC (A + C)", ctc],
#     ]

#     summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
#     summary_table.setStyle([
#         ("GRID", (0, 0), (-1, -1), 1, colors.black),
#         ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
#     ])

#     elements.append(Paragraph("<b>SUMMARY</b>", styles["Heading3"]))
#     elements.append(Spacer(1, 10))
#     elements.append(summary_table)

#     doc.build(elements)

#     pdf = buffer.getvalue()
#     buffer.close()
#     return pdf


# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def download_payslip_pdf(request, payslip_id):

#     try:
#         payslip = Payslip.objects.get(id=payslip_id)
#     except Payslip.DoesNotExist:
#         return Response({"error": "Payslip not found"}, status=404)

#     # Allow HR OR employee owner
#     if not request.user.is_staff and request.user != payslip.employee.user:
#         return Response({"error": "Permission denied"}, status=403)

#     pdf = generate_payslip_pdf(payslip)

#     return HttpResponse(pdf, content_type="application/pdf")


# ============================================================
# EMAIL SERVICE (Celery-Ready)
# ============================================================

def send_payslip_email_service(payslip):

    log = PayslipEmailLog.objects.create(
        payslip=payslip,
        email=payslip.employee.email,
        status="PENDING"
    )

    try:
        pdf = generate_payslip_pdf(payslip)

        email = EmailMessage(
            subject=f"Payslip - {payslip.month.strftime('%B %Y')}",
            body="Please find attached your payslip.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[payslip.employee.email],
        )

        email.attach("Payslip.pdf", pdf, "application/pdf")
        email.send()

        log.status = "SUCCESS"
        log.sent_at = timezone.now()
        log.save()

        return True

    except Exception as e:
        log.status = "FAILED"
        log.error_message = str(e)
        log.save()
        return False


@api_view(["POST"])
@permission_classes([IsHR])
def send_single_payslip_email(request):
    employee_id = request.data.get("employee_id")
    month_value = request.data.get("month")

    month_date = datetime.strptime(month_value, "%Y-%m").date()

    payslip = Payslip.objects.filter(
        employee__id=employee_id,
        month__year=month_date.year,
        month__month=month_date.month,
        status__in=["APPROVED", "PAID"]
    ).first()

    if not payslip:
        return Response({"error": "Approved or Paid payslip not found"}, status=404)

    success = send_payslip_email_service(payslip)
    if not success:
        return Response({"error": "Failed to send email. Check email logs or configuration."}, status=500)
    return Response({"message": "Email sent"})


@api_view(["POST"])
@permission_classes([IsHR])
def bulk_email_payslips(request):
    month_value = request.data.get("month")
    month_date = datetime.strptime(month_value, "%Y-%m").date()

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status__in=["APPROVED", "PAID"]
    )

    total = payslips.count()
    success = 0

    for p in payslips:
        if send_payslip_email_service(p):
            success += 1

    return Response({
        "total": total,
        "completed": success,
        "failed": total - success
    })



# ============================================================
# BULK APPROVE PAYSLIPS
# ============================================================

@api_view(["POST"])
@permission_classes([IsHR])
@transaction.atomic
def bulk_approve_payslips(request):

    month_value = request.data.get("month")

    if not month_value:
        return Response({"error": "month required"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m").date()
        year = month_date.year
        month = month_date.month
    except ValueError:
        return Response({"error": "Invalid month format (YYYY-MM)"}, status=400)

    # =========================================================
    # 1️⃣ APPROVE ALL DRAFT PAYSLIPS
    # =========================================================
    approved_now = Payslip.objects.filter(
        month__year=year,
        month__month=month,
        status="NOT PAID"
    ).update(status="APPROVED")

    # =========================================================
    # 2️⃣ GET ALL ACTIVE EMPLOYEES WITH SALARY STRUCTURE
    # =========================================================
    salaried_employees = Employee.objects.filter(
        is_active=True,
        salary__isnull=False
    )

    total_salaried = salaried_employees.count()

    # =========================================================
    # 3️⃣ COUNT APPROVED PAYSLIPS
    # =========================================================
    approved_count = Payslip.objects.filter(
        month__year=year,
        month__month=month,
        status="APPROVED"
    ).count()

    # =========================================================
    # 4️⃣ AUTO-CLOSE ONLY IF ALL APPROVED
    # =========================================================
    payroll_month, created = PayrollMonth.objects.get_or_create(
        year=year,
        month=month,
        defaults={"status": "OPEN"}
    )

    if total_salaried > 0 and approved_count == total_salaried:
        payroll_month.status = "CLOSED"
        payroll_month.save()

        return Response({
            "message": "All payslips approved. Payroll month CLOSED.",
            "approved_count": approved_now,
            "approved_now": approved_now,
            "total_salaried": total_salaried
        })

    return Response({
        "message": "Payslips approved but payroll month remains OPEN (pending approvals).",
        "approved_count": approved_now,
        "approved_now": approved_now,
        "total_salaried": total_salaried,
        "approved_total": approved_count
    })


# ============================================================
# PAYSLIP LISTING
# ============================================================

@api_view(["GET"])
@permission_classes([IsEmployee])
def my_payslips(request):

    employee = request.user.employee_profile

    payslips = Payslip.objects.filter(employee=employee)

    data = [
        {
            "id": p.id,
            "month": p.month,
            "net_pay": p.net_pay,
            "status": p.status,
        }
        for p in payslips
    ]

    return Response(data)


@api_view(["GET"])
@permission_classes([IsHR])
def all_payslips(request):

    payslips = Payslip.objects.select_related("employee").all()

    data = [
        {
            "id": p.id,
            "employee_name": f"{p.employee.first_name} {p.employee.last_name}",
            "month": p.month,
            "net_pay": p.net_pay,
            "status": p.status,
        }
        for p in payslips
    ]

    return Response(data)


# ============================================================
# DOWNLOAD ALL PAYSLIPS ZIP
# ============================================================

@api_view(["GET"])
@permission_classes([IsHR])
def download_all_payslips_zip(request):

    month_value = request.query_params.get("month")

    if not month_value:
        return Response({"error": "month required"}, status=400)

    month_date = datetime.strptime(month_value, "%Y-%m").date()

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status="APPROVED"
    )

    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w") as zip_file:
        for p in payslips:
            pdf = generate_payslip_pdf(p)
            zip_file.writestr(
                f"Payslip_{p.employee.employee_id}.pdf",
                pdf
            )

    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = "attachment; filename=payslips.zip"
    return response


# ============================================================
# EMAIL LOGS
# ============================================================

@api_view(["GET"])
@permission_classes([IsHR])
def payslip_email_logs(request, payslip_id):

    logs = PayslipEmailLog.objects.filter(payslip__id=payslip_id)

    data = [
        {
            "email": l.email,
            "status": l.status,
            "sent_at": l.sent_at,
            "error": l.error_message,
        }
        for l in logs
    ]

    return Response(data)


# ============================================================
# EMAIL DASHBOARD
# ============================================================

@api_view(["GET"])
@permission_classes([IsHR])
def email_dashboard(request):

    total = PayslipEmailLog.objects.count()
    success = PayslipEmailLog.objects.filter(status="SUCCESS").count()
    failed = PayslipEmailLog.objects.filter(status="FAILED").count()

    return Response({
        "total": total,
        "success": success,
        "failed": failed,
    })


# ============================================================
# BULK EMAIL PROGRESS (SYNC VERSION)
# ============================================================

@api_view(["GET"])
@permission_classes([IsHR])
def bulk_email_progress(request, batch_id):

    # For now simple sync logic
    return Response({
        "message": "Sync mode – no background tracking",
        "batch_id": batch_id
    })


# ============================================================
# PAYROLL DASHBOARD SUMMARY
# ============================================================

# ============================================================
# PAYROLL DASHBOARD SUMMARY (Aligned with Frontend)
# ============================================================

@api_view(["GET"])
@permission_classes([IsHR])
def payroll_dashboard_summary(request):
    from datetime import date, datetime
    from calendar import monthrange
    from apps.accounts.tenant_utils import get_current_company

    company = get_current_company(request)
    month_value = request.query_params.get("month")

    if not month_value:
        today = timezone.now()
        month_value = f"{today.year}-{str(today.month).zfill(2)}"

    try:
        month_date = datetime.strptime(month_value, "%Y-%m").date()
    except ValueError:
        return Response({"error": "Invalid month format (YYYY-MM)"}, status=400)

    _, last_day = monthrange(month_date.year, month_date.month)
    start_date = date(month_date.year, month_date.month, 1)
    end_date = date(month_date.year, month_date.month, last_day)

    payslips = Payslip.objects.filter(month__range=(start_date, end_date))
    if company is not None:
        payslips = payslips.filter(employee__user__company=company)

    total = payslips.count()
    draft = payslips.filter(status="NOT PAID").count()
    approved = payslips.filter(status="APPROVED").count()
    paid = payslips.filter(status="PAID").count()

    pm_kwargs = {"year": month_date.year, "month": month_date.month}
    if company and hasattr(PayrollMonth, "company"):
        pm_kwargs["company"] = company
    payroll_month = PayrollMonth.objects.filter(**pm_kwargs).first()

    status_value = payroll_month.status if payroll_month else "OPEN"

    return Response({
        "month": month_date.strftime("%B %Y"),
        "status": status_value,
        "total": total,
        "draft": draft,
        "approved": approved,
        "paid": paid,
    })

# ============================================================
# PAYROLL STATUS (Frontend Payroll Page Uses This)
# ============================================================

@api_view(["GET"])
@permission_classes([IsHR])
def payroll_status(request):

    month_value = request.query_params.get("month")
    filter_status = request.query_params.get("status", "ALL")

    if not month_value:
        return Response({"error": "month required"}, status=400)

    month_date = datetime.strptime(month_value, "%Y-%m").date()

    # ============================================
    # GET PAYROLL MONTH STATUS
    # ============================================

    payroll_month = PayrollMonth.objects.filter(
        year=month_date.year,
        month=month_date.month
    ).first()

    payroll_state = payroll_month.status if payroll_month else "OPEN"

    employees = Employee.objects.all()

    response_data = []

    for emp in employees:

        # ============================================
        # GET CURRENT SALARY
        # ============================================

        salary = get_current_salary(emp)
        salary_set = salary is not None

        # ============================================
        # PAYSLIP
        # ============================================

        payslip = Payslip.objects.filter(
            employee=emp,
            month__year=month_date.year,
            month__month=month_date.month
        ).first()

        if filter_status != "ALL":
            if not payslip or payslip.status != filter_status:
                continue

        # ============================================
        # SALARY CALCULATION
        # ============================================

        if salary:
            salary_totals = calculate_salary_totals(salary)
            gross_salary = salary_totals["gross_salary"]
            ctc = salary_totals["ctc"]
        else:
            salary_totals = None
            gross_salary = Decimal("0.00")
            ctc = Decimal("0.00")

        payslip_total_deductions = (
            to_decimal(payslip.fixed_deductions) + to_decimal(payslip.lop_deduction)
            if payslip
            else Decimal("0.00")
        )

        payslip_ctc = (
            to_decimal(payslip.gross_salary)
            + to_decimal(payslip.employer_pf)
            + to_decimal(payslip.employer_esi)
            + (salary_totals["gratuity"] if salary_totals else Decimal("0.00"))
        ) if payslip else ctc

        # ============================================
        # RESPONSE
        # ============================================

        response_data.append({

            "employee_id": emp.id,
            "employee_name": f"{emp.first_name} {emp.last_name}",
            "department": getattr(emp, "department", None) or "N/A",
            "account_number": getattr(emp, "account_number", None) or "",
            "ifsc": getattr(emp, "ifsc", None) or "",

            "salary_set": salary_set,

            "payslip_generated": bool(payslip),

            "payslip_status": ("NOT PAID" if payslip and payslip.status == "GENERATED" else (payslip.status if payslip else None)),

            "payslip_id": payslip.id if payslip else None,

            # Salary info
            "gross_salary": payslip.gross_salary if payslip else gross_salary,

            "lop_days": payslip.lop_days if payslip else Decimal("0.00"),

            "lop_deduction": payslip.lop_deduction if payslip else Decimal("0.00"),

            "total_deductions": (
                payslip_total_deductions
                if payslip
                else (salary_totals["total_deductions"] if salary_totals else Decimal("0.00"))
            ),

            "net_pay": payslip.net_pay if payslip else (salary_totals["net_salary"] if salary_totals else Decimal("0.00")),

            # CTC
            "ctc": payslip_ctc,
        })

    # ============================================
    # FINAL RESPONSE
    # ============================================

    return Response({
        "payroll_status": payroll_state,
        "month": month_value,
        "employees": response_data
    })

# ============================================================
# REOPEN PAYROLL MONTH
# ============================================================

@api_view(["POST"])
@permission_classes([IsHR])
def reopen_payroll_month(request):

    month_value = request.data.get("month")

    if not month_value:
        return Response({"error": "month required"}, status=400)

    month_date = datetime.strptime(month_value, "%Y-%m").date()

    PayrollMonth.objects.update_or_create(
        year=month_date.year,
        month=month_date.month,
        defaults={"status": "OPEN"}
    )

    return Response({"message": "Payroll month reopened"})


def calculate_professional_tax(gross_after_lop, employee):
    if not getattr(employee, 'pt_applicable', False):
        return Decimal("0.00")

    # Andhra Pradesh PT Slab (example)
    if gross_after_lop <= 15000:
        return Decimal("0.00")
    elif gross_after_lop <= 20000:
        return Decimal("150.00")
    else:
        return Decimal("200.00")



def calculate_leave_encashment(employee, year, month, gross_salary):

    total_days = monthrange(year, month)[1]

    encashable_balances = LeaveBalance.objects.filter(
        employee=employee,
        leave_type__encashable=True
    )

    total_encash_days = Decimal("0.0")

    for balance in encashable_balances:
        total_encash_days += balance.balance

    if total_encash_days <= 0:
        return Decimal("0.0"), Decimal("0.00")

    per_day_salary = Decimal(gross_salary) / Decimal(total_days)
    encash_amount = per_day_salary * total_encash_days

    return total_encash_days, encash_amount


@api_view(["GET"])
@permission_classes([IsHR])
def epf_report(request):

    month_value = request.query_params.get("month")

    if not month_value:
        return Response({"error": "month required (YYYY-MM)"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m")
    except ValueError:
        return Response({"error": "Invalid month format"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status__in=["APPROVED", "PAID"]
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="EPF_{month_value}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "UAN",
        "Employee Name",
        "Basic After LOP",
        "Employee PF",
        "Employer PF"
    ])

    for slip in payslips:
        writer.writerow([
            getattr(slip.employee, "uan_number", ""),
            slip.employee.first_name,
            slip.basic,
            slip.employee_pf,
            slip.employer_pf
        ])

    return response


@api_view(["GET"])
@permission_classes([IsHR])
def esi_report(request):

    month_value = request.query_params.get("month")

    if not month_value:
        return Response({"error": "month required (YYYY-MM)"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m")
    except ValueError:
        return Response({"error": "Invalid month format"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status__in=["APPROVED", "PAID"],
        employee_esi__gt=0
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="ESI_{month_value}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "IP Number",
        "Employee Name",
        "Gross After LOP",
        "Employee ESI",
        "Employer ESI"
    ])

    for slip in payslips:
        writer.writerow([
            getattr(slip.employee, "esi_number", ""),
            slip.employee.first_name,
            slip.gross_salary - slip.lop_deduction,
            slip.employee_esi,
            slip.employer_esi
        ])

    return response


@api_view(["GET"])
@permission_classes([IsHR])
def pt_report(request):

    month_value = request.query_params.get("month")

    if not month_value:
        return Response({"error": "month required (YYYY-MM)"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m")
    except ValueError:
        return Response({"error": "Invalid month format"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status__in=["APPROVED", "PAID"],
        professional_tax__gt=0
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="PT_{month_value}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Employee Name",
        "State",
        "Professional Tax"
    ])

    for slip in payslips:
        writer.writerow([
            slip.employee.first_name,
            getattr(slip.employee, "work_location", "") or getattr(settings, "COMPANY_STATE", "N/A"),
            slip.professional_tax
        ])

    return response


@api_view(["GET"])
@permission_classes([IsHR])
def epfo_ecr_file(request):

    month_value = request.query_params.get("month")

    if not month_value:
        return Response({"error": "month required (YYYY-MM)"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m")
    except ValueError:
        return Response({"error": "Invalid month format"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status__in=["APPROVED", "PAID"],
        employee_pf__gt=0
    ).select_related("employee")

    response = HttpResponse(content_type="text/plain")
    response["Content-Disposition"] = f'attachment; filename="ECR_{month_value}.txt"'

    for slip in payslips:

        employee = slip.employee
        # salary = employee.salary
        salary = get_current_salary(employee)

        # ===============================
        # EPF WAGES (Ceiling ₹15,000)
        # ===============================
        pf_wages = min(slip.basic, Decimal("15000"))

        # ===============================
        # EPS Calculation (8.33%)
        # ===============================
        eps_contribution = (pf_wages * Decimal("0.0833")).quantize(
            Decimal("1"), rounding=ROUND_HALF_UP
        )

        if eps_contribution > 1250:
            eps_contribution = Decimal("1250")

        # ===============================
        # EPF Contribution (Employee 12%)
        # ===============================
        epf_contribution = (pf_wages * Decimal("0.12")).quantize(
            Decimal("1"), rounding=ROUND_HALF_UP
        )

        # ===============================
        # NCP Days (LOP)
        # ===============================
        ncp_days = slip.lop_days or 0

        # ===============================
        # Construct ECR Row
        # ===============================
        row = [
            getattr(employee, "uan_number", ""),
            employee.first_name.upper(),
            str(int(slip.gross_salary)),
            str(int(pf_wages)),
            str(int(pf_wages)),  # EPS wages
            str(int(pf_wages)),  # EDLI wages
            str(int(epf_contribution)),
            str(int(eps_contribution)),
            str(int(ncp_days)),
            "0"  # Refund of advances
        ]

        response.write("|".join(row))
        response.write("\n")

    return response



@api_view(["GET"])
@permission_classes([IsHR])
def esic_upload_file(request):

    month_value = request.query_params.get("month")

    if not month_value:
        return Response({"error": "month required (YYYY-MM)"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m")
    except ValueError:
        return Response({"error": "Invalid month format"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status__in=["APPROVED", "PAID"],
        employee_esi__gt=0
    ).select_related("employee")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="ESIC_{month_value}.csv"'

    response.write("IP Number,Employee Name,Gross Wages,Employee Contribution,Employer Contribution\n")

    for slip in payslips:

        employee = slip.employee

        gross_after_lop = slip.gross_salary - slip.lop_deduction

        # Eligibility check (₹21,000 rule)
        if gross_after_lop > Decimal("21000"):
            continue

        # Employee 0.75%
        employee_esi = (gross_after_lop * Decimal("0.0075")).quantize(
            Decimal("1"), rounding=ROUND_HALF_UP
        )

        # Employer 3.25%
        employer_esi = (gross_after_lop * Decimal("0.0325")).quantize(
            Decimal("1"), rounding=ROUND_HALF_UP
        )

        row = [
            getattr(employee, "esi_number", ""),
            employee.first_name.upper(),
            str(int(gross_after_lop)),
            str(int(employee_esi)),
            str(int(employer_esi)),
        ]

        response.write(",".join(row))
        response.write("\n")

    return response


@api_view(["GET"])
@permission_classes([IsHR])
def generate_form16(request):

    employee_id = request.query_params.get("employee_id")
    financial_year = request.query_params.get("financial_year")  # 2025-2026

    if not employee_id or not financial_year:
        return Response(
            {"error": "employee_id and financial_year required"},
            status=400
        )

    try:
        start_year = int(financial_year.split("-")[0])
        end_year = start_year + 1
    except:
        return Response({"error": "Invalid financial year format"}, status=400)

    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    fy_start = date(start_year, 4, 1)
    fy_end = date(end_year, 3, 31)

    payslips = Payslip.objects.filter(
        employee=employee,
        month__gte=fy_start,
        month__lte=fy_end,
        status__in=["APPROVED", "PAID"]
    )

    total_gross = sum(p.gross_salary for p in payslips)
    total_tds = sum(p.tds_amount for p in payslips)
    total_net = sum(p.net_pay for p in payslips)

    # PDF Response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Form16_{financial_year}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()

    elements.append(Paragraph("<b>FORM 16</b>", styles["Title"]))
    elements.append(Spacer(1, 0.3 * inch))

    employer_name = getattr(settings, "COMPANY_NAME", "Your Company")
    employer_pan = getattr(settings, "COMPANY_PAN", "N/A")
    employer_tan = getattr(settings, "COMPANY_TAN", "N/A")

    # Employer Details
    employer_data = [
        ["Employer Name:", employer_name],
        ["Employer PAN:", employer_pan],
        ["Employer TAN:", employer_tan],
    ]

    table = Table(employer_data, colWidths=[2.5 * inch, 3 * inch])
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    # Employee Details
    employee_data = [
        ["Employee Name:", employee.first_name],
        ["Employee PAN:", employee.pan or "N/A"],
        ["Financial Year:", financial_year],
    ]

    table2 = Table(employee_data, colWidths=[2.5 * inch, 3 * inch])
    table2.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(table2)
    elements.append(Spacer(1, 0.3 * inch))

    # Salary Summary
    salary_data = [
        ["Total Gross Salary", f"₹ {total_gross:,.2f}"],
        ["Total TDS Deducted", f"₹ {total_tds:,.2f}"],
        ["Total Net Salary Paid", f"₹ {total_net:,.2f}"],
    ]

    table3 = Table(salary_data, colWidths=[3 * inch, 2.5 * inch])
    table3.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
    ]))

    elements.append(table3)

    doc.build(elements)

    return response


@api_view(["GET"])
@permission_classes([IsHR])
def generate_neft_file(request):

    month_value = request.query_params.get("month")

    if not month_value:
        return Response({"error": "month required (YYYY-MM)"}, status=400)

    try:
        month_date = datetime.strptime(month_value, "%Y-%m")
    except ValueError:
        return Response({"error": "Invalid month format"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=month_date.year,
        month__month=month_date.month,
        status="APPROVED"
    ).select_related("employee")

    if not payslips.exists():
        return Response({"error": "No approved payslips found"}, status=400)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="NEFT_{month_value}.csv"'

    writer = csv.writer(response)

    # Header
    writer.writerow([
        "Beneficiary Name",
        "Account Number",
        "IFSC Code",
        "Amount",
        "Payment Mode",
        "Remarks"
    ])

    total_amount = Decimal("0.00")

    for slip in payslips:
        employee = slip.employee

        if not getattr(employee, 'account_number', '') or not getattr(employee, 'ifsc', ''):
            continue  # Skip if bank details missing

        amount = slip.net_pay
        total_amount += amount

        writer.writerow([
            employee.first_name.upper(),
            getattr(employee, 'account_number', ''),
            getattr(employee, 'ifsc', ''),
            f"{amount:.2f}",
            "NEFT",
            f"Salary {month_value}"
        ])

    # Optional: summary row
    writer.writerow([])
    writer.writerow(["", "", "TOTAL", f"{total_amount:.2f}"])

    return response



@api_view(["GET"])
@permission_classes([IsHR])
def ctc_yearly_breakdown(request):

    employee_id = request.query_params.get("employee_id")

    if not employee_id:
        return Response({"error": "employee_id required"}, status=400)

    try:
        employee = Employee.objects.select_related("salary").get(id=employee_id)
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    if not hasattr(employee, "salary"):
        return Response({"error": "Salary structure not set"}, status=400)

    # salary = employee.salary
    salary = get_current_salary(employee)

    # Calculate salary components safely
    if salary:
        monthly_gross = (
            (getattr(salary, 'basic', 0) or 0) +
            (getattr(salary, 'da', 0) or 0) +
            (getattr(salary, 'hra', 0) or 0) +
            (getattr(salary, 'conveyance', 0) or 0) +
            (getattr(salary, 'medical', 0) or 0) +
            (getattr(salary, 'special_allowance', 0) or 0)
        )
        monthly_deductions = (
            (getattr(salary, 'employee_pf', 0) or 0) +
            (getattr(salary, 'professional_tax', 0) or 0) +
            (getattr(salary, 'employee_esi', 0) or 0) +
            (getattr(salary, 'tds', 0) or 0) +
            (getattr(salary, 'medical_insurance', 0) or 0)
        )
        monthly_net = monthly_gross - monthly_deductions
        monthly_employer = (
            (getattr(salary, 'employer_pf', 0) or 0) +
            (getattr(salary, 'employer_esi', 0) or 0) +
            (getattr(salary, 'gratuity', 0) or 0)
        )
        monthly_ctc = monthly_gross + monthly_employer
    else:
        monthly_gross = monthly_deductions = monthly_net = monthly_employer = monthly_ctc = 0

    # ================= YEARLY =================
    yearly_gross = monthly_gross * Decimal("12")
    yearly_deductions = monthly_deductions * Decimal("12")
    yearly_net = monthly_net * Decimal("12")
    yearly_employer = monthly_employer * Decimal("12")
    yearly_ctc = monthly_ctc * Decimal("12")

    return Response({
        "employee_name": f"{employee.first_name} {employee.last_name}",

        "monthly": {
            "gross_salary": monthly_gross,
            "total_deductions": monthly_deductions,
            "net_salary": monthly_net,
            "employer_contribution": monthly_employer,
            "ctc": monthly_ctc,
        },

        "yearly": {
            "gross_salary": yearly_gross,
            "total_deductions": yearly_deductions,
            "net_salary": yearly_net,
            "employer_contribution": yearly_employer,
            "ctc": yearly_ctc,
        },

        "breakdown_yearly": {
            "basic": salary.basic * 12,
            "da": salary.da * 12,
            "hra": salary.hra * 12,
            "conveyance": salary.conveyance * 12,
            "medical": salary.medical * 12,
            "special_allowance": salary.special_allowance * 12,
            "employee_pf": salary.employee_pf * 12,
            "professional_tax": salary.professional_tax * 12,
            "employee_esi": salary.employee_esi * 12,
            "tds": salary.tds * 12,
            "medical_insurance": salary.medical_insurance * 12,
            "employer_pf": salary.employer_pf * 12,
            "employer_esi": salary.employer_esi * 12,
            "gratuity": salary.gratuity * 12,
        }
    })


@api_view(["GET"])
@permission_classes([IsEmployee])
def my_payroll_summary(request):

    employee = request.user.employee_profile
    current_year = timezone.now().year

    records = Payslip.objects.filter(
        employee=employee,
        month__year=current_year
    )

    latest = records.order_by("-month").first()

    summary = {

        "employee_id": employee.id,

        "latest_net_pay": latest.net_pay if latest else 0,

        "ytd_earnings": records.aggregate(
            total=Sum("gross_salary")
        )["total"] or 0,

        "ytd_pf": records.aggregate(
            total=Sum("employee_pf")
        )["total"] or 0,

        "ytd_tax": records.aggregate(
            total=Sum("tds_amount")
        )["total"] or 0,

        "ytd_lop_days": records.aggregate(
            total=Sum("lop_days")
        )["total"] or 0,

    }

    return Response(summary)


@api_view(["GET"])
@permission_classes([IsEmployee])
def my_salary_view(request):
    """
    Employee can view ONLY their own salary payment history
    """
    employee = request.user.employee_profile
    current_year = timezone.now().year
    
    # Get all payroll records for this employee
    payslips = Payslip.objects.filter(
        employee=employee,
        status__in=["APPROVED", "PAID"]
    ).order_by("-month")
    
    # Calculate YTD summary
    ytd_payslips = payslips.filter(month__year=current_year)
    
    total_salary_ytd = ytd_payslips.aggregate(
        total=Sum("net_pay")
    )["total"] or 0
    
    # Get last payment
    last_payment = payslips.first()
    
    # Format payment data
    payments = []
    for payslip in payslips:
        payments.append({
            "id": payslip.id,
            "month": payslip.month.strftime("%B %Y"),
            "net_salary": float(payslip.net_pay),
            "payment_date": payslip.paid_on if payslip.paid_on else None,
            "status": payslip.status,
        })
    
    # Summary data
    summary = {
        "total_salary_ytd": float(total_salary_ytd),
        "last_payment_amount": float(last_payment.net_pay) if last_payment else 0,
        "last_payment_date": last_payment.paid_on if last_payment else None,
    }
    
    return Response({
        "payments": payments,
        "summary": summary
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payroll_summary(request):

    year = request.GET.get("year")
    month = request.GET.get("month")
    company_id = request.GET.get("company_id")

    payslips = Payslip.objects.all()

    if request.user.role == "SUPER_ADMIN":
        if company_id:
            payslips = payslips.filter(company_id=company_id)
    else:
        company = get_current_company(request)
        if company and hasattr(payslips.model, 'company'):
            payslips = payslips.filter(company=company)

    if year and month:
        payslips = payslips.filter(
            month__year=year,
            month__month=month
        )

    total_monthly_gross = payslips.aggregate(
        total=Sum("gross_salary")
    )["total"] or 0

    total_net_pay = payslips.aggregate(
        total=Sum("net_pay")
    )["total"] or 0

    total_employees = payslips.count()

    average_monthly_gross = (
        total_monthly_gross / total_employees
        if total_employees > 0 else 0
    )

    return Response({
        "total_monthly_ctc": total_monthly_gross,
        "total_monthly_gross": total_monthly_gross,
        "total_net_pay": total_net_pay,
        "total_employees": total_employees,
        "average_monthly_ctc": average_monthly_gross,
        "average_monthly_gross": average_monthly_gross,
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_payslip_pdf(request, payslip_id):

    try:
        payslip = Payslip.objects.select_related("employee").get(id=payslip_id)
    except Payslip.DoesNotExist:
        return Response({"error": "Payslip not found"}, status=404)

    allowed_roles = {"HR", "ADMIN", "SUPER_ADMIN"}
    user_role = getattr(request.user, "role", "")
    if user_role not in allowed_roles and request.user != payslip.employee.user:
        return Response({"error": "Permission denied"}, status=403)

    pdf = generate_payslip_pdf(payslip)

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Payslip_{payslip.month}.pdf"'

    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_payroll_excel(request):

    year = request.GET.get("year")
    month = request.GET.get("month")

    payslips = Payslip.objects.select_related("employee")

    if year and month:
        payslips = payslips.filter(
            month__year=year,
            month__month=month
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"

    headers = [
        "Employee ID",
        "Employee Name",
        "Month",
        "Gross Salary",
        "Net Pay"
    ]

    ws.append(headers)

    for slip in payslips:
        ws.append([
            slip.employee.employee_id,
            f"{slip.employee.first_name} {slip.employee.last_name}",
            slip.month.strftime("%B %Y"),
            float(slip.gross_salary),
            float(slip.net_pay),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="payroll_report.xlsx"'

    wb.save(response)
    return response



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_payroll_pdf(request):

    year = request.GET.get("year")
    month = request.GET.get("month")

    payslips = Payslip.objects.select_related("employee")

    if year and month:
        payslips = payslips.filter(
            month__year=year,
            month__month=month
        )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="payroll_report_{month}_{year}.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=pagesizes.A4)
    elements = []
    styles = getSampleStyleSheet()

    # ================= HEADER =================
    # Use current user's company if available, otherwise fallback
    company = getattr(request.user, "company", None)
    company_name = company.name if company else "Payroll Report"
    elements.append(Paragraph(f"<b>{company_name}</b>", styles["Title"]))
    elements.append(Spacer(1, 0.2 * inch))

    if year and month:
        month_name = datetime(
            int(year),
            int(month),
            1
        ).strftime("%B %Y")
        elements.append(
            Paragraph(
                f"<b>Payroll Report – {month_name}</b>",
                styles["Heading2"]
            )
        )
        elements.append(Spacer(1, 0.3 * inch))

    # ================= SUMMARY =================
    total_gross = payslips.aggregate(
        total=Sum("gross_salary")
    )["total"] or 0

    total_net = payslips.aggregate(
        total=Sum("net_pay")
    )["total"] or 0

    total_employees = payslips.count()

    summary_data = [
        ["Total Employees", total_employees],
        ["Total Gross Salary", f"₹ {total_gross:,.2f}"],
        ["Total Net Pay", f"₹ {total_net:,.2f}"],
    ]

    summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 0.5 * inch))

    # ================= EMPLOYEE TABLE =================

    data = [
        [
            "Employee ID",
            "Employee Name",
            "Gross Salary",
            "Net Pay"
        ]
    ]

    for slip in payslips:
        data.append([
            slip.employee.employee_id,
            f"{slip.employee.first_name} {slip.employee.last_name}",
            f"₹ {slip.gross_salary:,.2f}",
            f"₹ {slip.net_pay:,.2f}",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
    ]))

    elements.append(table)

    # ================= FOOTER =================
    elements.append(Spacer(1, 0.5 * inch))
    elements.append(
        Paragraph(
            "Generated by Genius Minds Making Code HRMS System",
            "Generated by Genius Minds Making Code Pvt Ltd HRMS System",
            styles["Normal"]
        )
    )

    doc.build(elements)

    return response



class SalaryRevisionViewSet(ModelViewSet):

    queryset = SalaryRevision.objects.all()
    serializer_class = SalaryRevisionSerializer
    permission_classes = [IsAuthenticated]


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_salary_revision(request):

    try:
        employee_id = request.data.get("employee")

        if not employee_id:
            return Response({"error": "Employee ID is required"}, status=400)

        employee = Employee.objects.get(id=employee_id)

        salary_components = normalize_salary_data(request.data)

        revision = SalaryRevision.objects.create(

            employee=employee,
            effective_from=request.data.get("effective_from"),

            reason=request.data.get("reason"),
            notes=request.data.get("notes"),

            basic=salary_components["basic"],
            da=salary_components["da"],
            hra=salary_components["hra"],
            conveyance=salary_components["conveyance"],
            medical=salary_components["medical"],
            special_allowance=salary_components["special_allowance"],
            employee_pf=salary_components["employee_pf"],
            professional_tax=salary_components["professional_tax"],
            employee_esi=salary_components["employee_esi"],
            tds=salary_components["tds"],
            medical_insurance=salary_components["medical_insurance"],
            employer_pf=salary_components["employer_pf"],
            employer_esi=salary_components["employer_esi"],
            gratuity=salary_components["gratuity"],
        )

        # Update current active salary structure
        try:
            from apps.employees.serializers import save_employee_salary
            save_employee_salary(employee, salary_components)
        except Exception as exc:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception("Failed to update active Salary record during revision for employee %s: %s", employee.id, exc)

        return Response({
            "message": "Salary revision created successfully",
            "revision_id": revision.id
        })

    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
def employee_salary_history(request, employee_id):

    revisions = SalaryRevision.objects.filter(
        employee_id=employee_id
    ).order_by("effective_from")

    data = []
    previous_ctc = Decimal("0.00")

    for rev in revisions:
        totals = calculate_salary_totals(rev)
        ctc = totals["ctc"]

        data.append({
            "id": rev.id,
            "effective_from": rev.effective_from,
            "gross_salary": totals["gross_salary"],
            "ctc": ctc,
            "previous_ctc": previous_ctc,
            "reason": rev.reason,
            "notes": rev.notes,
            "basic": rev.basic,
            "hra": rev.hra,
            "da": rev.da,
            "conveyance": rev.conveyance,
            "medical": rev.medical,
            "special_allowance": rev.special_allowance
        })
        previous_ctc = ctc

    return Response(data)


@api_view(["POST"])
@permission_classes([IsHR])
def export_salary_bank_file(request):
    month_value = request.data.get("month")
    year_value = request.data.get("year")
    company_account = request.data.get("company_account", "1234567890")

    if not month_value or not year_value:
        return Response({"error": "month and year required"}, status=400)

    try:
        month = int(month_value)
        year = int(year_value)
    except ValueError:
        return Response({"error": "Invalid month or year"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=year,
        month__month=month,
        status__in=["APPROVED", "PAID"],
        employee__is_active=True
    ).select_related("employee")

    if not payslips.exists():
        return Response({"error": "No approved payslips found"}, status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Salary Upload"

    headers = [
        "Debit Account Number",
        "Transaction Amount",
        "Transaction Currency",
        "Beneficiary Name",
        "Beneficiary Account Number",
        "Beneficiary IFSC Code",
        "Transaction Date",
        "Payment Mode",
        "Customer Reference Number",
        "Beneficiary Nickname/Code"
    ]
    ws.append(headers)

    transaction_date = datetime(year, month, 28).strftime("%d-%m-%Y")
    reference = f"Salary_{datetime(year, month, 1).strftime('%b_%Y')}"

    for slip in payslips:
        emp = slip.employee
        
        if not emp.account_number or not emp.ifsc:
            continue

        ws.append([
            company_account,
            float(slip.net_pay),
            "INR",
            f"{emp.first_name} {emp.last_name}".upper(),
            emp.account_number,
            emp.ifsc,
            transaction_date,
            "N",
            reference,
            emp.employee_id
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Bank_Salary_Upload_{reference}.xlsx"'

    wb.save(response)
    return response



@api_view(["GET"])
@permission_classes([IsHR])
def export_salary_bank_file(request):
    """
    Generate Excel file for bank salary transfer upload
    Query params: month, year, company_account (optional)
    """
    month_value = request.GET.get("month")
    year_value = request.GET.get("year")
    company_account = request.GET.get("company_account", "1234567890")

    if not month_value or not year_value:
        return Response({"error": "month and year required"}, status=400)

    try:
        month = int(month_value)
        year = int(year_value)
    except ValueError:
        return Response({"error": "Invalid month or year"}, status=400)

    payslips = Payslip.objects.filter(
        month__year=year,
        month__month=month,
        status__in=["APPROVED", "PAID"],
        employee__is_active=True
    ).select_related("employee")

    if not payslips.exists():
        return Response({"error": "No approved payslips found"}, status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Transfer"

    from openpyxl.styles import Font, Alignment, PatternFill

    headers = [
        "Debit Account Number",
        "Transaction Amount",
        "Transaction Currency",
        "Beneficiary Name",
        "Beneficiary Account Number",
        "Beneficiary IFSC Code",
        "Transaction Date",
        "Payment Mode",
        "Customer Reference Number",
        "Beneficiary Nickname/Code"
    ]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    transaction_date = datetime(year, month, 28).strftime("%d-%m-%Y")
    month_name = datetime(year, month, 1).strftime('%b')
    reference = f"Salary{month_name}{year}"

    row_num = 2

    for slip in payslips:
        emp = slip.employee
        
        if not emp.account_number or not emp.ifsc:
            continue

        ws.cell(row=row_num, column=1).value = company_account
        ws.cell(row=row_num, column=2).value = float(slip.net_pay)
        ws.cell(row=row_num, column=3).value = "INR"
        ws.cell(row=row_num, column=4).value = f"{emp.first_name} {emp.last_name}".upper()
        ws.cell(row=row_num, column=5).value = emp.account_number
        ws.cell(row=row_num, column=6).value = emp.ifsc
        ws.cell(row=row_num, column=7).value = transaction_date
        ws.cell(row=row_num, column=8).value = "NEFT"
        ws.cell(row=row_num, column=9).value = reference
        ws.cell(row=row_num, column=10).value = f"{emp.employee_id} - {emp.first_name} {emp.last_name}"
        row_num += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Bank_Salary_Transfer_{month_name}_{year}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
